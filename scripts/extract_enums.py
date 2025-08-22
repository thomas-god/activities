# /// script
# requires-python = ">=3.13"
# dependencies = [
#     "openpyxl",
# ]
# ///

from email.mime import base
from openpyxl import load_workbook


def parse_row_columns(row):
    type_name, base_type, value_name, value, _rest = row

    return (type_name.value, base_type.value, value_name.value, value.value)


def parse_type(iterator):
    mapping = {}
    next_type_name = None
    next_base_type = None

    for row in iterator:
        type_name, base_type, value_name, value = parse_row_columns(row)

        if type_name is not None and base_type is not None:
            next_type_name = type_name
            next_base_type = base_type
            break

        mapping[value] = value_name

    return (mapping, next_type_name, next_base_type)


def parse_enums(file):
    wb = load_workbook(filename=file, read_only=True)
    fit_types = wb["Types"]

    enums = []

    iterator = fit_types.iter_rows()
    next(iterator)  # Skip header

    type_name, base_type, _, _ = parse_row_columns(next(iterator))
    while True:
        mapping, next_type_name, next_base_type = parse_type(iterator)

        enums.append((type_name, base_type, mapping))

        if next_type_name is None and next_base_type is None:
            break

        type_name = next_type_name
        base_type = next_base_type
    return enums


def to_camel_case(string: str):
    if len(string) == 0:
        return string

    while string[0].isdigit():
        string = string[1:]

    if len(string) == 0:
        return string

    return "".join(w.capitalize() for w in string.lower().split("_"))


type_mapping = {
    "enum": "u8",
    "uint8": "u8",
    "uint8z": "u8",
    "uint16": "u16",
    "uint32": "u32",
    "uint32z": "u32",
}


def build_rust_enums(enums):
    content = ""

    # Defines enum of all other enums
    _enums = ",\n".join(
        f"{to_camel_case(e)}({to_camel_case(e)})" for (e, _, _) in enums
    )
    content += f"""
#[derive(Debug, PartialEq)]
pub enum Enums {{
{_enums}
}}
"""

    for enum, base_type, mapping in enums:
        if len(mapping) == 0:
            continue

        enum = to_camel_case(enum)
        variants = ",\n".join(to_camel_case(v) for v in mapping.values())
        mappings = ",\n".join(
            f"{k} => {enum}::{to_camel_case(v)}" for k, v in mapping.items()
        )
        # Define the enum
        content += f"""
#[derive(Debug, PartialEq)]
pub enum {enum} {{
{variants},
Unknown
}}"""

        # Define the mapping from byte to enum variant
        content += f"""
impl {enum} {{
pub fn from(content: {type_mapping[base_type]}) -> {enum} {{
match content {{
{mappings},
_ => {enum}::Unknown
}}
}}
}}
"""

    return content


if __name__ == "__main__":
    enums = parse_enums("Profile.xlsx")
    rust_enums = build_rust_enums(enums)

    print(rust_enums)
